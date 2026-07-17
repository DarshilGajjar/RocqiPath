"""
cell_counting.py
=================
Whole-slide DAB-positive cell quantification using an HSV brown-colour gate
+ OTSU thresholding on the inverted Value channel within that gate. Applies
to any DAB-chromogen IHC marker your dataset targets — the algorithm
itself is marker-agnostic.

Algorithm (per patch)
---------------------
1.  Brown colour gate in HSV:
        Hue  ∈ [5, 20]   (orange-brown — excludes haematoxylin H ≈ 110-140)
        Sat  ≥ 30        (excludes pale background / grey)
        Val  ≤ 220       (excludes white background)
2.  Invert the Value channel so dark-brown cells become bright.
3.  OTSU threshold on the inverted-Value pixels that fall inside the brown
    gate — computed SEPARATELY for GT and Prediction (user request).
4.  Connected-component labelling + min/max area filter.

Outputs
-------
  • JSON  — total counts, density (cells/mm²), positivity %, tissue area
  • Excel — per-patch GT vs Pred counts, Otsu thresholds, difference, summary stats
  • PNG   — dark-themed 2×3 panel per patch: Original | Brown gate | Binary overlay

References
----------
  Galon J. et al., Science 2006  (cells/mm² metric)
  Ruifrok & Johnston 2001        (H-DAB colour reference)
"""

import json
import os
import warnings
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import cv2
import numpy as np
import openslide
from skimage import filters, measure, morphology
from tqdm.auto import tqdm
from roqcipath.magnification import DEFAULT_TARGET_MAGNIFICATION
from roqcipath.output import OutputLayout
from roqcipath.slide import SlideReader as _SlideReader

warnings.filterwarnings("ignore")

# ── Aperio standard MPP (used when metadata is unavailable) ──────────────────
_APERIO_MPP = 0.2528   # µm/px at 40×, Aperio AT2

# ── WSI extensions ────────────────────────────────────────────────────────────
WSI_EXTENSIONS = frozenset(
    {".svs", ".tif", ".tiff", ".ome.tif", ".ome.tiff",
     ".ndpi", ".scn", ".mrxs", ".vms", ".vmu"}
)


# ── Slide reader with PIL fallback ────────────────────────────────────────────
# ── Excel output ──────────────────────────────────────────────────────────────
def _write_excel(results: list, output_path: str) -> None:
    """
    Write per-patch results and summary statistics to a two-sheet Excel workbook.
    Each row: patch name, GT count, Pred count, difference, GT OTSU, Pred OTSU.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL  = PatternFill("solid", start_color="2F5597")
    ALT_FILL  = PatternFill("solid", start_color="DCE6F1")
    SUM_FILL  = PatternFill("solid", start_color="FFC000")
    THIN      = Side(border_style="thin", color="BFBFBF")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BODY_FONT = Font(name="Arial", size=10)
    BOLD_FONT = Font(name="Arial", bold=True, size=10)
    CENTER    = Alignment(horizontal="center", vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center")

    def _cell(ws, row, col, value, font=None, fill=None,
              align=None, num_fmt=None):
        """Write one formatted cell into the worksheet.

        Parameters
        ----------
        ws : openpyxl.worksheet.worksheet.Worksheet
            The worksheet to write into.
        row, col : int
            1-based row and column indices for the target cell.
        value : Any
            The value to write into the cell.
        font : openpyxl.styles.Font, optional
            Font to apply. Defaults to ``BODY_FONT`` (from the enclosing
            scope) when omitted.
        fill : openpyxl.styles.PatternFill, optional
            Background fill to apply. Left as the worksheet default
            (no fill) when omitted.
        align : openpyxl.styles.Alignment, optional
            Text alignment to apply. Defaults to ``CENTER`` (from the
            enclosing scope) when omitted.
        num_fmt : str, optional
            Excel number format string (e.g. ``"0.00"``). Left as the
            default general format when omitted.

        Returns
        -------
        openpyxl.cell.cell.Cell
            The cell object that was written and formatted, in case the
            caller needs to make further adjustments.

        Notes
        -----
        A thin border (``BORDER``, from the enclosing scope) is always
        applied, regardless of the ``fill``/``font``/``align``
        arguments, so every cell in the sheet has a consistent grid
        appearance.
        """
        c = ws.cell(row=row, column=col, value=value)
        c.font      = font  or BODY_FONT
        c.alignment = align or CENTER
        c.border    = BORDER
        if fill:   c.fill          = fill
        if num_fmt: c.number_format = num_fmt
        return c

    wb = Workbook()

    # ── Sheet 1: per-patch counts ─────────────────────────────────────────────
    ws = wb.active
    ws.title        = "Cell Counts"
    ws.freeze_panes = "A2"

    headers = [
        "Patch Name",
        "GT Cell Count",
        "Pred Cell Count",
        "Difference (Pred−GT)",
        "GT OTSU threshold",
        "Pred OTSU threshold",
        "GT Image Path",
        "Pred Image Path",
    ]
    widths = [30, 18, 18, 22, 20, 22, 45, 45]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        _cell(ws, 1, col, h, font=HDR_FONT, fill=HDR_FILL)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    ds = 2  # data start row
    for i, r in enumerate(results):
        row  = ds + i
        fill = ALT_FILL if i % 2 == 0 else None
        gc, pc = r.get("gt_count"), r.get("pred_count")
        diff = (pc - gc) if (gc is not None and pc is not None) else None
        _cell(ws, row, 1, r.get("patch_name"),       fill=fill, align=LEFT)
        _cell(ws, row, 2, gc,                         fill=fill)
        _cell(ws, row, 3, pc,                         fill=fill)
        _cell(ws, row, 4, diff,                       fill=fill)
        _cell(ws, row, 5, r.get("gt_threshold"),      fill=fill, num_fmt="0.0")
        _cell(ws, row, 6, r.get("pred_threshold"),    fill=fill, num_fmt="0.0")
        _cell(ws, row, 7, str(r.get("gt_path")   or ""), fill=fill, align=LEFT)
        _cell(ws, row, 8, str(r.get("pred_path") or ""), fill=fill, align=LEFT)
        ws.row_dimensions[row].height = 18

    s = ds + len(results)
    _cell(ws, s, 1, "TOTAL / AVERAGE", font=BOLD_FONT, fill=SUM_FILL, align=LEFT)
    for col, formula in [
        (2, f"=SUM(B{ds}:B{s-1})"),
        (3, f"=SUM(C{ds}:C{s-1})"),
        (4, f"=SUM(D{ds}:D{s-1})"),
        (5, f"=AVERAGE(E{ds}:E{s-1})"),
        (6, f"=AVERAGE(F{ds}:F{s-1})"),
    ]:
        c = ws.cell(row=s, column=col, value=formula)
        c.font = BOLD_FONT; c.fill = SUM_FILL
        c.alignment = CENTER; c.border = BORDER
        if col in (5, 6): c.number_format = "0.0"
    ws.row_dimensions[s].height = 22

    # ── Sheet 2: summary statistics ───────────────────────────────────────────
    ws2 = wb.create_sheet("Summary Statistics")
    for col, (h, w) in enumerate(zip(
            ["Metric", "Ground Truth", "Predicted"], [32, 20, 20]), 1):
        _cell(ws2, 1, col, h, font=HDR_FONT, fill=HDR_FILL)
        ws2.column_dimensions[get_column_letter(col)].width = w

    def _stat(arr):
        """Compute the 7 summary statistics shown in the Summary Statistics sheet.

        Parameters
        ----------
        arr : list of number
            Per-patch values to summarise (e.g. ground-truth or
            predicted counts across all patches).

        Returns
        -------
        list
            ``[count, sum, mean, median, std_dev, min, max]`` as
            ``[int, float, float, float, float, float, float]``, matching
            the 7 rows of ``labels_s`` (from the enclosing scope) in
            order. Returns ``[None] * 7`` if ``arr`` is empty, so an
            empty column of dashes appears rather than raising on
            zero-length input (e.g. when no ground-truth counts were
            available for a batch).
        """
        if not arr: return [None] * 7
        a = np.array(arr)
        return [len(a), float(a.sum()), float(a.mean()), float(np.median(a)),
                float(a.std()), float(a.min()), float(a.max())]

    gt_vals   = [r["gt_count"]   for r in results if r.get("gt_count")   is not None]
    pred_vals = [r["pred_count"] for r in results if r.get("pred_count") is not None]
    labels_s  = ["Patches Processed", "Total Cells", "Mean / Patch",
                 "Median / Patch", "Std Dev", "Min Count", "Max Count"]

    for i, (lbl, gv, pv) in enumerate(zip(labels_s, _stat(gt_vals), _stat(pred_vals))):
        row  = i + 2
        fill = ALT_FILL if i % 2 == 0 else None
        _cell(ws2, row, 1, lbl,  fill=fill, align=LEFT)
        _cell(ws2, row, 2, round(gv, 2) if gv is not None else "N/A", fill=fill)
        _cell(ws2, row, 3, round(pv, 2) if pv is not None else "N/A", fill=fill)

    wb.save(str(output_path))
    print(f"[INFO]  Excel saved → {output_path}")


# ── Main class ────────────────────────────────────────────────────────────────
class PositiveCellCounter:
    """Count DAB-positive cells across whole-slide images using an HSV
    brown-colour gate combined with per-image OTSU thresholding.

    Works with any DAB-chromogen IHC marker — the detection method
    itself has no notion of which biomarker produced the brown signal,
    only that it is brown. See the module docstring for the full
    algorithm description and references.

    Typical usage
    -------------
    ::

        counter = PositiveCellCounter({
            "patch_size":   512,
            "magnification": 2,
            "output_dir":   "./results/cell_counts",
        })
        result = counter.count_slide("./slide_01.svs")

    Parameters (cfg dict)
    ---------------------
    patch_size        : tile size in pixels at the chosen magnification (default 512)
    tissue_threshold  : minimum tissue fraction per patch (default 0.10)
    target_magnification : physical analysis zoom — default 20x
    output_dir        : root output folder
    min_cell_area     : minimum cell area in px² (default 50)
    max_cell_area     : maximum cell area in px², None = no upper bound
    """

    def __init__(self, cfg: dict):
        """Resolve configuration, create the output directory, and print a summary.

        Parameters
        ----------
        cfg : dict
            Configuration dict. All keys are optional:

            - ``"patch_size"`` (int) — tile edge length in pixels at the
              chosen magnification. Defaults to ``512``.
            - ``"tissue_threshold"`` (float) — minimum fraction of
              non-background pixels (see :meth:`_is_tissue`) for a patch
              to be processed at all. Defaults to ``0.10``.
            - ``"target_magnification"`` (float) — exact physical zoom for
              analysis. Defaults to ``20.0``. The legacy ``"magnification"``
              key is accepted as a physical-value alias.
            - ``"output_dir"`` (str) — root directory for results.
              Defaults to ``"./cell_count_output"``; created if it
              doesn't exist.
            - ``"min_cell_area"`` (int) — minimum connected-component
              area, in pixels², for a detected blob to count as a cell.
              Defaults to ``50``.
            - ``"max_cell_area"`` (int or None) — maximum connected-component
              area, in pixels². When omitted, empty, or ``0``, treated
              as "no upper bound" (``self.max_cell_area`` becomes
              ``None``).

        Notes
        -----
        Prints a startup summary (patch size, tissue threshold, cell
        area range, thresholding strategy) to stdout after resolving all
        fields.
        """
        self.patch_size       = int(cfg.get("patch_size",       512))
        self.tissue_threshold = float(cfg.get("tissue_threshold", 0.10))
        self.target_magnification = float(
            cfg.get("target_magnification", cfg.get("magnification", DEFAULT_TARGET_MAGNIFICATION))
        )
        self.source_magnification = cfg.get("source_magnification")
        self.paired_source_magnification = cfg.get("paired_source_magnification")
        self.output_dir       = os.path.abspath(cfg.get("output_dir", "./cell_count_output"))
        self.min_cell_area    = int(cfg.get("min_cell_area",    50))
        _max = cfg.get("max_cell_area", None)
        self.max_cell_area    = int(_max) if _max not in (None, "", 0) else None

        if self.patch_size <= 0:
            raise ValueError("patch_size must be > 0")
        if self.target_magnification <= 0:
            raise ValueError("target_magnification must be > 0")
        if not (0.0 <= self.tissue_threshold <= 1.0):
            raise ValueError("tissue_threshold must be in [0, 1]")
        if self.min_cell_area <= 0:
            raise ValueError("min_cell_area must be > 0")
        if self.max_cell_area is not None and self.max_cell_area < self.min_cell_area:
            raise ValueError("max_cell_area must be >= min_cell_area")

        self.layout = OutputLayout(self.output_dir)
        self.layout.module_dir("cell_counting")
        _max_str = f"{self.max_cell_area}" if self.max_cell_area else "∞"
        print("[INFO] Positive Cell Counter (HSV brown gate + OTSU)")
        print(f"       Patch size       : {self.patch_size} px at {self.target_magnification:g}x")
        print(f"       Tissue threshold : {int(self.tissue_threshold * 100)}%")
        print(f"       Cell area range  : {self.min_cell_area} – {_max_str} px²")
        print("       Thresholding     : OTSU computed separately per image")

    # ── Tissue gate ───────────────────────────────────────────────────────────
    def _is_tissue(self, rgb: np.ndarray) -> bool:
        """Decide whether a patch contains enough tissue to bother counting cells.

        Parameters
        ----------
        rgb : numpy.ndarray
            An RGB patch array (any integer dtype; mean brightness is
            computed across the colour channels).

        Returns
        -------
        bool
            ``True`` if the fraction of pixels whose mean RGB value is
            below 235 (i.e. not near-white background) is at least
            ``self.tissue_threshold``; ``False`` otherwise, meaning the
            patch is mostly blank slide background.
        """
        return float(np.mean(np.mean(rgb, axis=2) < 235)) >= self.tissue_threshold

    # ── MPP ───────────────────────────────────────────────────────────────────
    @staticmethod
    def _get_mpp(slide) -> Tuple[float, float]:
        """Read the microns-per-pixel (MPP) calibration from slide metadata.

        Parameters
        ----------
        slide : openslide.OpenSlide or similar
            An open slide handle exposing an OpenSlide-style
            ``.properties`` mapping.

        Returns
        -------
        tuple of (float, float)
            ``(mpp_x, mpp_y)`` read from the slide's
            ``openslide.mpp-x`` / ``openslide.mpp-y`` properties. Returns
            ``(0.0, 0.0)`` if the properties are missing or cannot be
            parsed as floats (caught via a broad ``except Exception``) —
            callers (e.g. :meth:`count_slide`) treat this as "MPP
            unavailable" and typically substitute a fallback constant
            (e.g. the Aperio standard) rather than failing outright.
        """
        try:
            props = slide.properties
            return (float(props[openslide.PROPERTY_NAME_MPP_X]),
                    float(props[openslide.PROPERTY_NAME_MPP_Y]))
        except Exception:
            return 0.0, 0.0

    # ── Core counting: HSV brown gate + OTSU ─────────────────────────────────
    @staticmethod
    def _brown_mask(img_rgb: np.ndarray) -> np.ndarray:
        """
        Boolean mask for brown (DAB-positive) pixels.
        Hue ∈ [5, 20], Sat ≥ 30, Val ≤ 220  (OpenCV HSV: H ∈ [0,180]).
        Explicitly excludes haematoxylin blue (H ≈ 110-140) and background.
        """
        hsv = cv2.cvtColor(img_rgb, cv2.COLOR_RGB2HSV)
        H, S, V = hsv[..., 0], hsv[..., 1], hsv[..., 2]
        return (H >= 5) & (H <= 20) & (S >= 30) & (V <= 220)

    def _count_patch(self, img_rgb: np.ndarray,
                     threshold: Optional[float] = None) -> Tuple:
        """
        Count positive cells in one RGB patch using HSV brown gate + OTSU.

        Parameters
        ----------
        img_rgb   : H × W × 3 uint8 RGB array
        threshold : if None → compute OTSU from this patch's own brown pixels
                    if float → apply this fixed threshold (for shared-threshold mode)

        Returns
        -------
        count, binary_mask, brown_vis, threshold_used, labels
        """
        brown = self._brown_mask(img_rgb)
        empty = np.zeros(img_rgb.shape[:2], bool)

        if brown.sum() < 10:
            return 0, empty, img_rgb.copy(), 0.0, np.zeros(img_rgb.shape[:2], int)

        # Invert Value so dark-brown cells become bright
        hsv     = cv2.cvtColor(img_rgb, cv2.COLOR_RGB2HSV)
        inv_val = cv2.bitwise_not(hsv[..., 2])

        # OTSU on brown pixels only (or use supplied fixed threshold)
        if threshold is None:
            threshold = float(filters.threshold_otsu(inv_val[brown]))

        binary = (inv_val > threshold) & brown
        binary = morphology.remove_small_objects(binary, min_size=self.min_cell_area)
        binary = morphology.remove_small_holes(binary,
                                               area_threshold=max(1, self.min_cell_area // 2))

        labels  = measure.label(binary)
        regions = measure.regionprops(labels)
        max_a   = self.max_cell_area or float("inf")
        valid   = [r.label for r in regions if self.min_cell_area <= r.area <= max_a]
        binary  = np.isin(labels, valid)
        labels  = measure.label(binary)
        count   = int(labels.max())

        # Visualisation: brown pixels in original colour, rest grey
        brown_vis = np.full_like(img_rgb, 210)
        brown_vis[brown] = img_rgb[brown]

        return count, binary, brown_vis, float(threshold), labels

    # ── Patch comparison plot ─────────────────────────────────────────────────
    def _save_comparison_plot(
        self,
        gt_rgb: np.ndarray,   gt_result: Tuple,
        pred_rgb: np.ndarray, pred_result: Tuple,
        patch_idx: int, x: int, y: int,
        save_path: str, dpi: int = 130,
    ) -> None:
        """
        Dark-themed 2×3 panel:
          Row 1 (GT)  : Original | Brown gate | Binary overlay
          Row 2 (Pred): Original | Brown gate | Binary overlay
        """
        import matplotlib.pyplot as plt

        gt_count,   gt_bin,   gt_brown,   gt_thr,   _ = gt_result
        pred_count, pred_bin, pred_brown, pred_thr, _ = pred_result

        diff = pred_count - gt_count
        pct  = f"{diff/gt_count*100:+.1f}%" if gt_count > 0 else "N/A"

        def _overlay(rgb, binary):
            """Paint detected-cell pixels bright green over the original RGB image.

            Parameters
            ----------
            rgb : numpy.ndarray
                The original RGB patch, ``(H, W, 3)``.
            binary : numpy.ndarray
                A boolean mask, same ``(H, W)`` shape as ``rgb``'s first
                two dimensions, ``True`` where a detected cell pixel is.

            Returns
            -------
            numpy.ndarray
                A copy of ``rgb`` with every pixel where ``binary`` is
                ``True`` recoloured to bright green (``[30, 210, 30]``),
                for visual QC of the detection overlay.
            """
            ov = rgb.copy()
            ov[binary] = [30, 210, 30]
            return ov

        rows = [
            (gt_rgb,   gt_brown,   _overlay(gt_rgb,   gt_bin),
             f"GROUND TRUTH\nCells: {gt_count}   OTSU θ={gt_thr:.1f}",
             "#2ecc71"),
            (pred_rgb, pred_brown, _overlay(pred_rgb, pred_bin),
             f"PREDICTED\nCells: {pred_count}   OTSU θ={pred_thr:.1f}",
             "#e74c3c"),
        ]
        col_titles = [
            "Original Image",
            "Brown Colour Gate\n(grey = excluded)",
            "Binary Detection\n(green = positive cell)",
        ]

        fig, axes = plt.subplots(2, 3, figsize=(14, 9))
        fig.patch.set_facecolor("#1a1a2e")
        fig.suptitle(
            f"Patch {patch_idx:04d}  ·  x={x}  y={y}"
            f"     GT: {gt_count}   Pred: {pred_count}   Δ = {diff:+d}  ({pct})",
            fontsize=12, fontweight="bold", color="white", y=1.01,
        )

        for r, (orig, brown_v, overlay, row_lbl, colour) in enumerate(rows):
            for c, (ax, panel) in enumerate(zip(axes[r], [orig, brown_v, overlay])):
                ax.set_facecolor("#1a1a2e")
                ax.imshow(panel)
                if r == 0:
                    ax.set_title(col_titles[c], fontsize=9,
                                 fontweight="bold", color="white", pad=6)
                if c == 0:
                    ax.set_ylabel(row_lbl, fontsize=8,
                                  color=colour, fontweight="bold", labelpad=6)
                for spine in ax.spines.values():
                    spine.set_edgecolor(colour)
                    spine.set_linewidth(1.5)
                ax.tick_params(left=False, bottom=False,
                               labelleft=False, labelbottom=False)

        plt.tight_layout()
        plt.savefig(save_path, dpi=dpi, bbox_inches="tight",
                    facecolor="#1a1a2e")
        plt.close(fig)

    # ── Single slide ──────────────────────────────────────────────────────────
    def count_slide(self, wsi_path: str, label: str = "Cell") -> dict:
        """Count DAB+ positive cells across one slide. Each patch uses its own OTSU."""
        slide      = _SlideReader(wsi_path)
        plan = slide.configure_magnification(
            self.target_magnification, self.source_magnification
        )
        w, h = slide.target_dimensions
        mpp_x, mpp_y = self._get_mpp(slide)
        slide_name = Path(wsi_path).stem

        print(f"[INFO] Slide       : {Path(wsi_path).name}")
        print(f"[INFO] Dimensions  : {w} × {h} px")
        if mpp_x:
            print(f"[INFO] Resolution  : {mpp_x:.4f} µm/px  (from metadata)")
            mpp_x *= plan.level0_per_target_pixel
            mpp_y *= plan.level0_per_target_pixel
        else:
            level0_mpp = _APERIO_MPP
            mpp_x = mpp_y = level0_mpp * plan.level0_per_target_pixel
            print(f"[WARN] MPP not in metadata — using {mpp_x:.4f} µm/px at target zoom")

        total_pos = total_nuc = tissue_px = 0
        tiles_x = (w + self.patch_size - 1) // self.patch_size
        tiles_y = (h + self.patch_size - 1) // self.patch_size

        with tqdm(total=tiles_x * tiles_y,
                  desc=f"  {slide_name}", unit="patch") as pbar:
            for py in range(0, h, self.patch_size):
                for px in range(0, w, self.patch_size):
                    tw = min(self.patch_size, w - px)
                    th = min(self.patch_size, h - py)
                    patch = slide.read_at_magnification((px, py), (tw, th)).convert("RGB")
                    rgb = np.array(patch)
                    patch.close()

                    if self._is_tissue(rgb):
                        count, *_ = self._count_patch(rgb)  # own OTSU
                        total_pos += count
                        tissue_px += tw * th
                    pbar.update(1)

        slide.close()

        px_mm2          = (mpp_x / 1000.0) * (mpp_y / 1000.0)
        tissue_area_mm2 = tissue_px * px_mm2
        density_per_mm2 = total_pos / tissue_area_mm2 if tissue_area_mm2 > 0 else 0.0

        print("\n[RESULT] ══════════════════════════════════")
        print(f"[RESULT]  DAB+ cells    : {total_pos:,}")
        print(f"[RESULT]  Tissue area   : {tissue_area_mm2:.3f} mm²")
        print(f"[RESULT]  Density       : {density_per_mm2:.1f} cells/mm²")
        print("[RESULT] ══════════════════════════════════")

        results = {
            "slide":            Path(wsi_path).name,
            "label":            label,
            "total_positive":   int(total_pos),
            "tissue_area_mm2":  round(tissue_area_mm2, 4),
            "density_per_mm2":  round(density_per_mm2, 2),
            "mpp_x": mpp_x, "mpp_y": mpp_y,
        }
        out_dir = self.layout.item_dir("cell_counting", slide_name)
        json_path = out_dir / f"{slide_name}_cell_count_results.json"
        with open(json_path, "w") as f:
            json.dump(results, f, indent=2)
        print(f"[INFO]  Saved → {json_path}")
        return results

    # ── GT vs Prediction pair ─────────────────────────────────────────────────
    def count_slide_pair(
        self,
        gt_path:    str,
        pred_path:  str,
        label:      str  = "Cell",
        save_plots: bool = True,
        max_plots:  int  = 10,
        dpi:        int  = 150,
    ) -> dict:
        """
        Count DAB+ positive cells in a GT slide and a Prediction slide.

        OTSU threshold is computed INDEPENDENTLY for each image —
        both GT and Pred derive their own threshold from their own
        brown-pixel distribution.

        Saves:
          • Per-patch comparison PNGs  (dark 2×3 panel, up to max_plots)
          • JSON summary
          • Excel workbook with per-patch counts + statistics
        """
        import matplotlib
        matplotlib.use("Agg")

        gt_slide   = _SlideReader(gt_path)
        pred_slide = _SlideReader(pred_path)
        gt_plan = gt_slide.configure_magnification(
            self.target_magnification, self.source_magnification
        )
        pred_slide.configure_magnification(
            self.target_magnification, self.paired_source_magnification
        )
        w, h = gt_slide.target_dimensions
        if pred_slide.target_dimensions != (w, h):
            raise ValueError(
                f"Slides differ at {self.target_magnification:g}x: "
                f"{w}x{h} vs {pred_slide.target_dimensions[0]}x{pred_slide.target_dimensions[1]}"
            )
        mpp_x, mpp_y = self._get_mpp(gt_slide)

        print(f"[INFO] GT slide    : {Path(gt_path).name}")
        print(f"[INFO] Pred slide  : {Path(pred_path).name}")
        print(f"[INFO] Dimensions  : {w} × {h} px")
        if mpp_x:
            print(f"[INFO] Resolution  : {mpp_x:.4f} µm/px  (from metadata)")
            mpp_x *= gt_plan.level0_per_target_pixel
            mpp_y *= gt_plan.level0_per_target_pixel
        else:
            mpp_x = mpp_y = _APERIO_MPP * gt_plan.level0_per_target_pixel
            print(f"[WARN] MPP not in metadata — using {mpp_x:.4f} µm/px at target zoom")

        item_name = f"{Path(gt_path).stem}_vs_{Path(pred_path).stem}"
        out_dir = self.layout.item_dir("cell_counting", item_name)
        plots_dir = out_dir

        gt_total = pred_total = tissue_px = 0
        plots_saved = patch_idx = 0
        patch_results: List[dict] = []

        tiles_x = (w + self.patch_size - 1) // self.patch_size
        tiles_y = (h + self.patch_size - 1) // self.patch_size

        with tqdm(total=tiles_x * tiles_y,
                  desc="  GT vs Pred", unit="patch") as pbar:
            for py in range(0, h, self.patch_size):
                for px in range(0, w, self.patch_size):
                    tw = min(self.patch_size, w - px)
                    th = min(self.patch_size, h - py)

                    gt_patch = gt_slide.read_at_magnification((px, py), (tw, th)).convert("RGB")
                    gt_rgb   = np.array(gt_patch); gt_patch.close()

                    if not self._is_tissue(gt_rgb):
                        pbar.update(1)
                        continue

                    patch_idx += 1
                    tissue_px += tw * th

                    pred_patch = pred_slide.read_at_magnification((px, py), (tw, th)).convert("RGB")
                    pred_rgb   = np.array(pred_patch); pred_patch.close()

                    # ── Independent OTSU per image ────────────────────────────
                    gt_result   = self._count_patch(gt_rgb,   threshold=None)
                    pred_result = self._count_patch(pred_rgb, threshold=None)

                    g_count, _, _, g_thr, _ = gt_result
                    p_count, _, _, p_thr, _ = pred_result

                    gt_total   += g_count
                    pred_total += p_count

                    patch_results.append({
                        "patch_name":    f"patch_{patch_idx:04d}",
                        "gt_count":      g_count,
                        "pred_count":    p_count,
                        "gt_threshold":  round(g_thr, 1),
                        "pred_threshold": round(p_thr, 1),
                        "gt_path":       str(gt_path),
                        "pred_path":     str(pred_path),
                    })

                    # ── Save comparison plot ───────────────────────────────────
                    if save_plots and plots_saved < max_plots:
                        fname = str(plots_dir /
                                    f"patch_{patch_idx:04d}_x{px}_y{py}.png")
                        self._save_comparison_plot(
                            gt_rgb, gt_result, pred_rgb, pred_result,
                            patch_idx, px, py, fname, dpi=dpi,
                        )
                        plots_saved += 1
                        tqdm.write(
                            f"  [PLOT] {plots_saved}/{max_plots}  "
                            f"GT={g_count} (θ={g_thr:.1f})  "
                            f"Pred={p_count} (θ={p_thr:.1f})"
                        )

                    pbar.update(1)

        gt_slide.close()
        pred_slide.close()

        # ── Metrics ───────────────────────────────────────────────────────────
        px_mm2          = (mpp_x / 1000.0) * (mpp_y / 1000.0)
        tissue_area_mm2 = tissue_px * px_mm2
        density_gt      = gt_total   / tissue_area_mm2 if tissue_area_mm2 > 0 else 0.0
        density_pred    = pred_total / tissue_area_mm2 if tissue_area_mm2 > 0 else 0.0
        diff_abs        = pred_total - gt_total
        diff_pct        = (diff_abs / gt_total * 100) if gt_total > 0 else float("nan")

        print("\n[RESULT] ══════════════════════════════════════════════")
        print(f"[RESULT]  Tissue patches    : {patch_idx}")
        print(f"[RESULT]  GT  DAB+ cells   : {gt_total:,}")
        print(f"[RESULT]  Pred DAB+ cells  : {pred_total:,}")
        print(f"[RESULT]  Δ absolute        : {diff_abs:+,}")
        if gt_total > 0:
            print(f"[RESULT]  Δ relative        : {diff_pct:+.1f}%")
        print(f"[RESULT]  Tissue area       : {tissue_area_mm2:.3f} mm²")
        print(f"[RESULT]  GT  density       : {density_gt:.1f} cells/mm²")
        print(f"[RESULT]  Pred density      : {density_pred:.1f} cells/mm²")
        if save_plots:
            print(f"[RESULT]  Plots saved       : {plots_saved}  →  {plots_dir}")
        print("[RESULT] ══════════════════════════════════════════════")

        summary = {
            "gt_slide":              Path(gt_path).name,
            "pred_slide":            Path(pred_path).name,
            "label":                 label,
            "gt_positive":           int(gt_total),
            "pred_positive":         int(pred_total),
            "diff_absolute":         int(diff_abs),
            "diff_pct":              round(diff_pct, 2) if gt_total > 0 else None,
            "tissue_area_mm2":       round(tissue_area_mm2, 4),
            "gt_density_per_mm2":    round(density_gt,   2),
            "pred_density_per_mm2":  round(density_pred, 2),
            "plots_saved":           plots_saved,
            "thresholding":          "independent_otsu_per_image",
        }

        # JSON
        json_path = out_dir / (
            f"{Path(gt_path).stem}_vs_{Path(pred_path).stem}_results.json")
        with open(json_path, "w") as f:
            json.dump(summary, f, indent=2)
        print(f"[INFO]  JSON  → {json_path}")

        # Excel
        excel_path = out_dir / (
            f"{Path(gt_path).stem}_vs_{Path(pred_path).stem}_counts.xlsx")
        _write_excel(patch_results, str(excel_path))

        return summary

    # ── Batch ─────────────────────────────────────────────────────────────────
    def count_batch(self, input_dir: str, label: str = "Cell") -> list:
        """Process all WSI files in input_dir (single-slide mode)."""
        slides = sorted(
            p for p in Path(input_dir).iterdir()
            if any(str(p).lower().endswith(ext) for ext in WSI_EXTENSIONS)
        )
        if not slides:
            print(f"[ERROR] No WSI files found in: {input_dir}")
            return []

        print(f"[INFO] Found {len(slides)} slide(s)")
        all_results = []
        for slide_path in slides:
            print(f"\n{'─'*50}")
            try:
                result = self.count_slide(str(slide_path), label=label)
                all_results.append(result)
            except Exception as e:
                print(f"[ERROR] {slide_path.name}: {e}")

        if all_results:
            tot = sum(r["total_positive"] for r in all_results)
            print("\n[BATCH] ══════════════════════════════════")
            print(f"[BATCH]  Slides  : {len(all_results)}")
            print(f"[BATCH]  Total   : {tot:,} DAB+ cells")
            print("[BATCH] ══════════════════════════════════")

            out = self.layout.module_dir("cell_counting")
            with open(out / "batch_cell_count_results.json", "w") as f:
                json.dump(all_results, f, indent=2)

        return all_results
