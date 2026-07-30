"""Excel and comparison-figure writers for cell-counting results."""

from __future__ import annotations

from typing import Tuple

import numpy as np


def _write_excel(results: list, output_path: str) -> None:
    """Write per-patch results and summary statistics to an Excel workbook.

    Each row: patch name, GT count, Pred count, difference, GT OTSU, Pred OTSU.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL = PatternFill("solid", start_color="2F5597")
    ALT_FILL = PatternFill("solid", start_color="DCE6F1")
    SUM_FILL = PatternFill("solid", start_color="FFC000")
    THIN = Side(border_style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BODY_FONT = Font(name="Arial", size=10)
    BOLD_FONT = Font(name="Arial", bold=True, size=10)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center")

    def _cell(ws, row, col, value, font=None, fill=None, align=None, num_fmt=None):
        """Write one formatted cell into the worksheet.

        Parameters
        ----------
        ws : openpyxl.worksheet.worksheet.Worksheet
            The worksheet to write into.
        row, col : int
            1-based row and column indices for the target cell.
        value : Any
            The value to write into the cell.
        font : openpyxl.styles.Font, optional
            Font to apply. Defaults to ``BODY_FONT`` (from the enclosing
            scope) when omitted.
        fill : openpyxl.styles.PatternFill, optional
            Background fill to apply. Left as the worksheet default
            (no fill) when omitted.
        align : openpyxl.styles.Alignment, optional
            Text alignment to apply. Defaults to ``CENTER`` (from the
            enclosing scope) when omitted.
        num_fmt : str, optional
            Excel number format string (e.g. ``"0.00"``). Left as the
            default general format when omitted.

        Returns
        -------
        openpyxl.cell.cell.Cell
            The cell object that was written and formatted, in case the
            caller needs to make further adjustments.

        Notes
        -----
        A thin border (``BORDER``, from the enclosing scope) is always
        applied, regardless of the ``fill``/``font``/``align``
        arguments, so every cell in the sheet has a consistent grid
        appearance.
        """
        c = ws.cell(row=row, column=col, value=value)
        c.font = font or BODY_FONT
        c.alignment = align or CENTER
        c.border = BORDER
        if fill:
            c.fill = fill
        if num_fmt:
            c.number_format = num_fmt
        return c

    wb = Workbook()

    # ── Sheet 1: per-patch counts ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Cell Counts"
    ws.freeze_panes = "A2"

    headers = [
        "Patch Name",
        "GT Cell Count",
        "Pred Cell Count",
        "Difference (Pred−GT)",
        "GT OTSU threshold",
        "Pred OTSU threshold",
        "GT Image Path",
        "Pred Image Path",
    ]
    widths = [30, 18, 18, 22, 20, 22, 45, 45]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        _cell(ws, 1, col, h, font=HDR_FONT, fill=HDR_FILL)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    ds = 2  # data start row
    for i, r in enumerate(results):
        row = ds + i
        fill = ALT_FILL if i % 2 == 0 else None
        gc, pc = r.get("gt_count"), r.get("pred_count")
        diff = (pc - gc) if (gc is not None and pc is not None) else None
        _cell(ws, row, 1, r.get("patch_name"), fill=fill, align=LEFT)
        _cell(ws, row, 2, gc, fill=fill)
        _cell(ws, row, 3, pc, fill=fill)
        _cell(ws, row, 4, diff, fill=fill)
        _cell(ws, row, 5, r.get("gt_threshold"), fill=fill, num_fmt="0.0")
        _cell(ws, row, 6, r.get("pred_threshold"), fill=fill, num_fmt="0.0")
        _cell(ws, row, 7, str(r.get("gt_path") or ""), fill=fill, align=LEFT)
        _cell(ws, row, 8, str(r.get("pred_path") or ""), fill=fill, align=LEFT)
        ws.row_dimensions[row].height = 18

    s = ds + len(results)
    _cell(ws, s, 1, "TOTAL / AVERAGE", font=BOLD_FONT, fill=SUM_FILL, align=LEFT)
    for col, formula in [
        (2, f"=SUM(B{ds}:B{s - 1})"),
        (3, f"=SUM(C{ds}:C{s - 1})"),
        (4, f"=SUM(D{ds}:D{s - 1})"),
        (5, f"=AVERAGE(E{ds}:E{s - 1})"),
        (6, f"=AVERAGE(F{ds}:F{s - 1})"),
    ]:
        c = ws.cell(row=s, column=col, value=formula)
        c.font = BOLD_FONT
        c.fill = SUM_FILL
        c.alignment = CENTER
        c.border = BORDER
        if col in (5, 6):
            c.number_format = "0.0"
    ws.row_dimensions[s].height = 22

    # ── Sheet 2: summary statistics ───────────────────────────────────────────
    ws2 = wb.create_sheet("Summary Statistics")
    for col, (h, w) in enumerate(zip(["Metric", "Ground Truth", "Predicted"], [32, 20, 20]), 1):
        _cell(ws2, 1, col, h, font=HDR_FONT, fill=HDR_FILL)
        ws2.column_dimensions[get_column_letter(col)].width = w

    def _stat(arr):
        """Compute the 7 summary statistics shown in the Summary Statistics sheet.

        Parameters
        ----------
        arr : list of number
            Per-patch values to summarise (e.g. ground-truth or
            predicted counts across all patches).

        Returns
        -------
        list
            ``[count, sum, mean, median, std_dev, min, max]`` as
            ``[int, float, float, float, float, float, float]``, matching
            the 7 rows of ``labels_s`` (from the enclosing scope) in
            order. Returns ``[None] * 7`` if ``arr`` is empty, so an
            empty column of dashes appears rather than raising on
            zero-length input (e.g. when no ground-truth counts were
            available for a batch).
        """
        if not arr:
            return [None] * 7
        a = np.array(arr)
        return [
            len(a),
            float(a.sum()),
            float(a.mean()),
            float(np.median(a)),
            float(a.std()),
            float(a.min()),
            float(a.max()),
        ]

    gt_vals = [r["gt_count"] for r in results if r.get("gt_count") is not None]
    pred_vals = [r["pred_count"] for r in results if r.get("pred_count") is not None]
    labels_s = [
        "Patches Processed",
        "Total Cells",
        "Mean / Patch",
        "Median / Patch",
        "Std Dev",
        "Min Count",
        "Max Count",
    ]

    for i, (lbl, gv, pv) in enumerate(zip(labels_s, _stat(gt_vals), _stat(pred_vals))):
        row = i + 2
        fill = ALT_FILL if i % 2 == 0 else None
        _cell(ws2, row, 1, lbl, fill=fill, align=LEFT)
        _cell(ws2, row, 2, round(gv, 2) if gv is not None else "N/A", fill=fill)
        _cell(ws2, row, 3, round(pv, 2) if pv is not None else "N/A", fill=fill)

    wb.save(str(output_path))
    print(f"[INFO]  Excel saved → {output_path}")


def _save_comparison_plot(
    gt_rgb: np.ndarray,
    gt_result: Tuple,
    pred_rgb: np.ndarray,
    pred_result: Tuple,
    patch_idx: int,
    x: int,
    y: int,
    save_path: str,
    dpi: int = 130,
) -> None:
    """Save a dark-themed 2×3 comparison panel.

    Row 1 (GT)  : Original | Brown gate | Binary overlay
    Row 2 (Pred): Original | Brown gate | Binary overlay
    """
    import matplotlib.pyplot as plt

    gt_count, gt_bin, gt_brown, gt_thr, _ = gt_result
    pred_count, pred_bin, pred_brown, pred_thr, _ = pred_result

    diff = pred_count - gt_count
    pct = f"{diff / gt_count * 100:+.1f}%" if gt_count > 0 else "N/A"

    def _overlay(rgb, binary):
        """Paint detected-cell pixels bright green over the original RGB image.

        Parameters
        ----------
        rgb : numpy.ndarray
            The original RGB patch, ``(H, W, 3)``.
        binary : numpy.ndarray
            A boolean mask, same ``(H, W)`` shape as ``rgb``'s first
            two dimensions, ``True`` where a detected cell pixel is.

        Returns
        -------
        numpy.ndarray
            A copy of ``rgb`` with every pixel where ``binary`` is
            ``True`` recoloured to bright green (``[30, 210, 30]``),
            for visual QC of the detection overlay.
        """
        ov = rgb.copy()
        ov[binary] = [30, 210, 30]
        return ov

    rows = [
        (
            gt_rgb,
            gt_brown,
            _overlay(gt_rgb, gt_bin),
            f"GROUND TRUTH\nCells: {gt_count}   OTSU θ={gt_thr:.1f}",
            "#2ecc71",
        ),
        (
            pred_rgb,
            pred_brown,
            _overlay(pred_rgb, pred_bin),
            f"PREDICTED\nCells: {pred_count}   OTSU θ={pred_thr:.1f}",
            "#e74c3c",
        ),
    ]
    col_titles = [
        "Original Image",
        "Brown Colour Gate\n(grey = excluded)",
        "Binary Detection\n(green = positive cell)",
    ]

    fig, axes = plt.subplots(2, 3, figsize=(14, 9))
    fig.patch.set_facecolor("#1a1a2e")
    fig.suptitle(
        f"Patch {patch_idx:04d}  ·  x={x}  y={y}"
        f"     GT: {gt_count}   Pred: {pred_count}   Δ = {diff:+d}  ({pct})",
        fontsize=12,
        fontweight="bold",
        color="white",
        y=1.01,
    )

    for r, (orig, brown_v, overlay, row_lbl, colour) in enumerate(rows):
        for c, (ax, panel) in enumerate(zip(axes[r], [orig, brown_v, overlay])):
            ax.set_facecolor("#1a1a2e")
            ax.imshow(panel)
            if r == 0:
                ax.set_title(col_titles[c], fontsize=9, fontweight="bold", color="white", pad=6)
            if c == 0:
                ax.set_ylabel(row_lbl, fontsize=8, color=colour, fontweight="bold", labelpad=6)
            for spine in ax.spines.values():
                spine.set_edgecolor(colour)
                spine.set_linewidth(1.5)
            ax.tick_params(left=False, bottom=False, labelleft=False, labelbottom=False)

    plt.tight_layout()
    plt.savefig(save_path, dpi=dpi, bbox_inches="tight", facecolor="#1a1a2e")
    plt.close(fig)
